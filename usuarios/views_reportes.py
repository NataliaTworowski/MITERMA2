from django.db.models import Sum, Count, Avg, Q
from django.db.models.functions import TruncDate
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import json
import csv
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect, render
from .decorators import admin_terma_required

# Para exportación Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Para exportación PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import io


@admin_terma_required
def reportes_premium(request):
    """Vista para mostrar reportes premium para administradores de termas."""
    usuario = request.user
    terma = usuario.terma
    
    # Verificar que el usuario tenga plan premium (a través de la terma)
    if not (hasattr(usuario, 'terma') and usuario.terma and usuario.terma.plan_actual and usuario.terma.plan_actual.nombre == 'premium'):
        messages.error(request, 'Esta funcionalidad requiere un plan premium.')
        return redirect('usuarios:adm_termas')
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    reportes_data = None
    
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            # Importar modelo Compra
            from ventas.models import Compra
            
            # Obtener compras del período
            compras = Compra.objects.filter(
                terma=terma,
                fecha_compra__date__gte=fecha_inicio_dt,
                fecha_compra__date__lte=fecha_fin_dt,
                estado_pago='pagado'
            )
            
            # Calcular KPIs
            total_ingresos = compras.aggregate(total=Sum('total'))['total'] or 0
            total_entradas = compras.aggregate(total=Sum('cantidad'))['total'] or 0
            total_clientes = compras.values('usuario').distinct().count()
            promedio_venta = compras.aggregate(avg=Avg('total'))['avg'] or 0
            
            # Nuevas métricas: Clientes nuevos vs recurrentes
            clientes_periodo = set(compras.values_list('usuario', flat=True).distinct())
            
            # Obtener compras anteriores al período para identificar clientes recurrentes
            compras_anteriores = Compra.objects.filter(
                terma=terma,
                fecha_compra__date__lt=fecha_inicio_dt,
                estado_pago='pagado'
            ).values_list('usuario', flat=True).distinct()
            
            clientes_anteriores = set(compras_anteriores)
            clientes_nuevos = len(clientes_periodo - clientes_anteriores)
            clientes_recurrentes = len(clientes_periodo & clientes_anteriores)
            
            # Porcentaje de retención
            if total_clientes > 0:
                porcentaje_retencion = (clientes_recurrentes / total_clientes) * 100
            else:
                porcentaje_retencion = 0
            
            # Comparativa de ventas con período anterior
            dias_periodo = (fecha_fin_dt - fecha_inicio_dt).days + 1
            fecha_inicio_anterior = fecha_inicio_dt - timedelta(days=dias_periodo)
            fecha_fin_anterior = fecha_inicio_dt - timedelta(days=1)
            
            ventas_anterior = Compra.objects.filter(
                terma=terma,
                fecha_compra__date__gte=fecha_inicio_anterior,
                fecha_compra__date__lte=fecha_fin_anterior,
                estado_pago='pagado'
            ).aggregate(total=Sum('total'))['total'] or 0
            
            # Calcular porcentaje de crecimiento
            if ventas_anterior > 0:
                crecimiento_ventas = ((total_ingresos - ventas_anterior) / ventas_anterior) * 100
            else:
                crecimiento_ventas = 100 if total_ingresos > 0 else 0
                
            crecimiento_positivo = crecimiento_ventas >= 0
            
            # Ventas por día
            ventas_por_dia = compras.annotate(
                fecha=TruncDate('fecha_compra')
            ).values('fecha').annotate(
                total_dia=Sum('total')
            ).order_by('fecha')
            
            # Tipos de entrada más vendidos - agrupar por nombre para evitar duplicados
            from entradas.models import EntradaTipo
            from django.db.models import Q
            from collections import defaultdict
            
            # Obtener datos agrupados por nombre de entrada
            tipos_vendidos = defaultdict(int)
            for compra in compras.prefetch_related('detalles__entrada_tipo'):
                for detalle in compra.detalles.all():
                    tipos_vendidos[detalle.entrada_tipo.nombre] += detalle.cantidad
            
            # Convertir a lista ordenada y tomar los top 5
            tipos_entrada_ordenados = sorted(tipos_vendidos.items(), key=lambda x: x[1], reverse=True)[:5]
            
            # Preparar datos para gráficos
            fechas_labels = []
            ventas_valores = []
            fecha_actual = fecha_inicio_dt
            while fecha_actual <= fecha_fin_dt:
                fechas_labels.append(fecha_actual.strftime('%d/%m'))
                venta_dia = next((v['total_dia'] for v in ventas_por_dia if v['fecha'] == fecha_actual), 0)
                ventas_valores.append(float(venta_dia))
                fecha_actual += timedelta(days=1)
            
            tipos_entrada_labels = [nombre for nombre, cantidad in tipos_entrada_ordenados]
            tipos_entrada_data = [cantidad for nombre, cantidad in tipos_entrada_ordenados]
            
            reportes_data = {
                'total_ingresos': total_ingresos,
                'total_entradas': total_entradas,
                'total_clientes': total_clientes,
                'promedio_venta': promedio_venta,
                'clientes_nuevos': clientes_nuevos,
                'clientes_recurrentes': clientes_recurrentes,
                'porcentaje_retencion': porcentaje_retencion,
                'crecimiento_ventas': crecimiento_ventas,
                'crecimiento_positivo': crecimiento_positivo,
                'ventas_anterior': ventas_anterior,
                'fechas_labels': fechas_labels,
                'ventas_por_dia': ventas_valores,
                'tipos_entrada_labels': tipos_entrada_labels,
                'tipos_entrada_data': tipos_entrada_data,
            }
            
            # Paginación para las ventas detalle
            compras_detalle = compras.select_related('usuario').prefetch_related('detalles__entrada_tipo')
            paginator = Paginator(compras_detalle, 10)  # 10 registros por página
            page_number = request.GET.get('page')
            ventas_paginadas = paginator.get_page(page_number)
            
            reportes_data['ventas_detalle'] = ventas_paginadas
            reportes_data['total_ventas_count'] = compras_detalle.count()
            
        except ValueError:
            messages.error(request, 'Formato de fecha inválido.')
    
    context = {
        'title': 'Reportes Premium - MiTerma',
        'usuario': usuario,
        'terma': terma,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'reportes_data': reportes_data,
    }
    
    return render(request, 'administrador_termas/reporte_premium.html', context)


@admin_terma_required
def exportar_reporte_csv(request):
    """Vista para exportar reportes a CSV."""
    from django.http import JsonResponse
    
    usuario = request.user
    terma = usuario.terma
    
    # Verificar que el usuario tenga plan premium (a través de la terma)
    if not (hasattr(usuario, 'terma') and usuario.terma and usuario.terma.plan_actual and usuario.terma.plan_actual.nombre == 'premium'):
        return JsonResponse({'error': 'Esta funcionalidad requiere un plan premium.'}, status=403)
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if not fecha_inicio or not fecha_fin:
        return JsonResponse({'error': 'Debe especificar fechas de inicio y fin.'}, status=400)
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Importar modelo Compra
        from ventas.models import Compra
        
        # Obtener compras del período
        compras = Compra.objects.filter(
            terma=terma,
            fecha_compra__date__gte=fecha_inicio_dt,
            fecha_compra__date__lte=fecha_fin_dt,
            estado_pago='pagado'
        ).select_related('usuario').prefetch_related('detalles__entrada_tipo')
        
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="reporte_{terma.nombre_terma}_{fecha_inicio}_{fecha_fin}.csv"'
        response.write('\ufeff')  # BOM para UTF-8
        
        writer = csv.writer(response)
        writer.writerow([
            'Fecha Compra', 'Nombre Cliente', 'Apellido Cliente', 'Email Cliente', 
            'Tipo Entrada', 'Cantidad', 'Precio Unitario', 'Subtotal', 'Total Compra'
        ])
        
        for compra in compras:
            for detalle in compra.detalles.all():
                writer.writerow([
                    compra.fecha_compra.strftime('%d/%m/%Y %H:%M'),
                    compra.usuario.nombre,
                    compra.usuario.apellido,
                    compra.usuario.email,
                    detalle.entrada_tipo.nombre,
                    detalle.cantidad,
                    detalle.precio_unitario,
                    detalle.subtotal,
                    compra.total
                ])
        
        return response
        
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error al generar CSV: {str(e)}'}, status=500)


@admin_terma_required
def exportar_reporte_excel(request):
    """Vista para exportar reportes a Excel."""
    from django.http import JsonResponse
    
    usuario = request.user
    terma = usuario.terma
    
    # Verificar que el usuario tenga plan premium
    if not (hasattr(usuario, 'terma') and usuario.terma and usuario.terma.plan_actual and usuario.terma.plan_actual.nombre == 'premium'):
        return JsonResponse({'error': 'Esta funcionalidad requiere un plan premium.'}, status=403)
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if not fecha_inicio or not fecha_fin:
        return JsonResponse({'error': 'Debe especificar fechas de inicio y fin.'}, status=400)
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        from ventas.models import Compra
        
        # Obtener compras del período
        compras = Compra.objects.filter(
            terma=terma,
            fecha_compra__date__gte=fecha_inicio_dt,
            fecha_compra__date__lte=fecha_fin_dt,
            estado_pago='pagado'
        ).select_related('usuario').prefetch_related('detalles__entrada_tipo')
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título del reporte
        ws.merge_cells('A1:I3')
        title_cell = ws['A1']
        title_cell.value = f"Reporte de Ventas - {terma.nombre_terma}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Información del período
        ws.merge_cells('A4:I4')
        period_cell = ws['A4']
        period_cell.value = f"Período: {fecha_inicio_dt.strftime('%d/%m/%Y')} - {fecha_fin_dt.strftime('%d/%m/%Y')}"
        period_cell.font = Font(size=12)
        period_cell.alignment = Alignment(horizontal="center")
        
        # Encabezados
        headers = [
            'Fecha Compra', 'Nombre Cliente', 'Apellido Cliente', 'Email Cliente',
            'Tipo Entrada', 'Cantidad', 'Precio Unitario', 'Subtotal', 'Total Compra'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Datos
        row = 7
        total_general = 0
        total_entradas = 0
        
        for compra in compras:
            for detalle in compra.detalles.all():
                ws.cell(row=row, column=1).value = compra.fecha_compra.strftime('%d/%m/%Y %H:%M')
                ws.cell(row=row, column=2).value = compra.usuario.nombre
                ws.cell(row=row, column=3).value = compra.usuario.apellido
                ws.cell(row=row, column=4).value = compra.usuario.email
                ws.cell(row=row, column=5).value = detalle.entrada_tipo.nombre
                ws.cell(row=row, column=6).value = detalle.cantidad
                ws.cell(row=row, column=7).value = float(detalle.precio_unitario)
                ws.cell(row=row, column=8).value = float(detalle.subtotal)
                ws.cell(row=row, column=9).value = float(compra.total)
                
                total_entradas += detalle.cantidad
                
                # Aplicar bordes
                for col in range(1, 10):
                    ws.cell(row=row, column=col).border = border
                
                row += 1
            
            total_general += float(compra.total)
        
        # Totales
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        total_cell = ws[f'A{row}']
        total_cell.value = "TOTALES:"
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="right")
        
        ws.cell(row=row, column=6).value = total_entradas
        ws.cell(row=row, column=6).font = Font(bold=True)
        
        ws.merge_cells(f'G{row}:H{row}')
        ws.cell(row=row, column=7).value = "TOTAL GENERAL:"
        ws.cell(row=row, column=7).font = Font(bold=True)
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
        
        ws.cell(row=row, column=9).value = total_general
        ws.cell(row=row, column=9).font = Font(bold=True, color="008000")
        
        # Ajustar ancho de columnas
        column_widths = [20, 15, 15, 25, 20, 10, 15, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_{terma.nombre_terma}_{fecha_inicio}_{fecha_fin}.xlsx"'
        
        wb.save(response)
        return response
        
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error al generar Excel: {str(e)}'}, status=500)


@admin_terma_required
def exportar_reporte_pdf(request):
    """Vista para exportar reportes a PDF."""
    from django.http import JsonResponse
    
    usuario = request.user
    terma = usuario.terma
    
    # Verificar que el usuario tenga plan premium
    if not (hasattr(usuario, 'terma') and usuario.terma and usuario.terma.plan_actual and usuario.terma.plan_actual.nombre == 'premium'):
        return JsonResponse({'error': 'Esta funcionalidad requiere un plan premium.'}, status=403)
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if not fecha_inicio or not fecha_fin:
        return JsonResponse({'error': 'Debe especificar fechas de inicio y fin.'}, status=400)
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        from ventas.models import Compra
        
        # Obtener compras del período
        compras = Compra.objects.filter(
            terma=terma,
            fecha_compra__date__gte=fecha_inicio_dt,
            fecha_compra__date__lte=fecha_fin_dt,
            estado_pago='pagado'
        ).select_related('usuario').prefetch_related('detalles__entrada_tipo')
        
        # Calcular totales
        total_ingresos = compras.aggregate(total=Sum('total'))['total'] or 0
        total_entradas = 0
        total_compras = compras.count()
        
        # Crear PDF en memoria
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=1  # Centrado
        )
        
        # Título
        elements.append(Paragraph(f"Reporte de Ventas - {terma.nombre_terma}", title_style))
        elements.append(Paragraph(
            f"Período: {fecha_inicio_dt.strftime('%d/%m/%Y')} - {fecha_fin_dt.strftime('%d/%m/%Y')}", 
            subtitle_style
        ))
        
        # KPIs principales
        kpi_data = [
            ['Concepto', 'Valor', 'Concepto', 'Valor'],
            ['Total Ingresos', f'${total_ingresos:,.2f}', 'Total Compras', str(total_compras)],
            ['Total Entradas Vendidas', '', 'Promedio por Venta', ''],
        ]
        
        # Calcular total de entradas
        for compra in compras:
            for detalle in compra.detalles.all():
                total_entradas += detalle.cantidad
        
        promedio_venta = total_ingresos / total_compras if total_compras > 0 else 0
        
        kpi_data[2][1] = str(total_entradas)
        kpi_data[2][3] = f'${promedio_venta:,.2f}'
        
        # Agregar título del resumen ejecutivo antes de la tabla
        elements.append(Paragraph("Resumen Ejecutivo", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        kpi_table = Table(kpi_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(kpi_table)
        elements.append(Spacer(1, 20))
        
        # Tabla de detalles
        elements.append(Paragraph("Detalle de Ventas", styles['Heading2']))
        
        # Encabezados de la tabla de detalles
        detail_data = [
            ['Fecha', 'Cliente', 'Tipo Entrada', 'Cant.', 'P. Unit.', 'Subtotal', 'Total']
        ]
        
        # Datos de la tabla
        for compra in compras:
            for detalle in compra.detalles.all():
                detail_data.append([
                    compra.fecha_compra.strftime('%d/%m/%Y'),
                    f"{compra.usuario.nombre} {compra.usuario.apellido}",
                    detalle.entrada_tipo.nombre,
                    str(detalle.cantidad),
                    f'${detalle.precio_unitario:,.0f}',
                    f'${detalle.subtotal:,.0f}',
                    f'${compra.total:,.0f}'
                ])
        
        # Fila de totales
        detail_data.append([
            'TOTALES', '', '', str(total_entradas), '', '', f'${total_ingresos:,.0f}'
        ])
        
        detail_table = Table(detail_data, colWidths=[
            0.8*inch, 1.5*inch, 1.2*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.8*inch
        ])
        
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(detail_table)
        
        # Construir PDF
        doc.build(elements)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{terma.nombre_terma}_{fecha_inicio}_{fecha_fin}.pdf"'
        
        return response
        
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error al generar PDF: {str(e)}'}, status=500)